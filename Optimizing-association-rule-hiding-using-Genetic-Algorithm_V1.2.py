import random
import math
from itertools import combinations
import pandas as pd
import ast
import openpyxl
import time
import os
# Example Transactions
# Define the itemset of 15 items
# items = [
#     'Bread', 'Milk', 'Eggs', 'Diapers', 'Tea', 'Cola', 'Butter', 
#     'Syrup', 'Trash Cans', 'Pancake Mix', 'Water'
# ]

# # Generate 200 random transactions
# transactions = []
# for _ in range(200):
#     transaction_size = random.randint(2, 9)  # Random transaction size between 2 and 5 items
#     transaction = set(random.sample(items, transaction_size))
#     transactions.append(transaction)

transactions=[

        {'Diapers', 'Milk'}, {'Milk', 'Water', 'Eggs', 'Trash Cans', 'Tea'}, {'Diapers', 'Cola', 'Water', 'Eggs', 'Butter', 'Trash Cans', 'Tea'}, {'Trash Cans', 'Cola', 'Syrup'}, {'Milk', 'Cola', 'Water', 'Bread', 'Eggs', 'Butter', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Eggs', 'Trash Cans'}, {'Milk', 'Cola', 'Bread', 'Butter', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Eggs', 'Trash Cans', 'Tea'}, {'Milk', 'Cola', 'Water', 'Bread', 'Butter', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Milk', 'Cola', 'Water', 'Bread', 'Eggs', 'Butter', 'Trash Cans', 'Tea'}, {'Pancake Mix', 'Diapers', 'Water', 'Eggs', 'Butter', 'Syrup'}, {'Tea', 'Butter', 'Cola'}, {'Pancake Mix', 'Milk', 'Cola', 'Water', 'Bread', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Butter', 'Water'}, {'Pancake Mix', 'Milk', 'Cola', 'Bread', 'Eggs', 'Butter', 'Tea'}, {'Pancake Mix', 'Water'}, {'Pancake Mix', 'Water', 'Cola', 'Bread'}, {'Tea', 'Diapers', 'Water', 'Syrup'}, {'Pancake Mix', 'Milk', 'Diapers', 'Cola', 'Water', 'Bread', 'Eggs', 'Butter', 'Trash Cans'}, {'Diapers', 'Syrup'}, {'Pancake Mix', 'Milk', 'Cola', 'Water', 'Butter', 'Tea'}, {'Pancake Mix', 'Milk', 'Water', 'Bread', 'Butter', 'Eggs', 'Trash Cans', 'Tea'}, {'Milk', 'Cola', 'Water', 'Butter', 'Trash Cans'}, {'Pancake Mix', 'Diapers', 'Cola', 'Milk', 'Water', 'Eggs', 'Butter', 'Tea', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Milk', 'Water', 'Eggs', 'Tea'}, {'Milk', 'Bread', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Milk', 'Cola', 'Bread', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Bread', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Milk', 'Water', 'Bread', 'Eggs', 'Butter', 'Tea', 'Syrup'}, {'Pancake Mix', 'Milk', 'Cola', 'Water', 'Bread', 'Eggs', 'Butter', 'Trash Cans', 'Syrup'}, {'Diapers', 'Syrup'}, {'Diapers', 'Cola', 'Milk', 'Bread', 'Eggs', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Eggs', 'Milk', 'Tea'}, {'Pancake Mix', 'Milk', 'Cola', 'Bread', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Butter'}, {'Milk', 'Tea', 'Bread'}, {'Pancake Mix', 'Diapers', 'Cola', 'Milk', 'Water', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Milk', 'Cola', 'Bread', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Trash Cans', 'Diapers'}, {'Cola', 'Bread'}, {'Pancake Mix', 'Milk', 'Water', 'Syrup'}, {'Diapers', 'Milk', 'Tea', 'Butter'}, {'Diapers', 'Milk', 'Bread'}, {'Trash Cans', 'Syrup'}, {'Tea', 'Bread'}, {'Pancake Mix', 'Eggs', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Trash Cans', 'Diapers'}, {'Butter', 'Milk', 'Tea', 'Syrup'}, {'Butter', 'Milk', 'Cola'}, {'Pancake Mix', 'Diapers', 'Milk', 'Water', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Milk', 'Cola', 'Bread'}, {'Syrup', 'Bread'}, {'Water', 'Syrup'}, {'Pancake Mix', 'Water'}, {'Pancake Mix', 'Trash Cans', 'Milk', 'Cola'}, {'Pancake Mix', 'Milk', 'Water', 'Bread', 'Eggs', 'Butter'}, {'Diapers', 'Cola', 'Milk', 'Bread', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Milk', 'Water', 'Tea', 'Syrup'}, {'Diapers', 'Cola', 'Milk', 'Water', 'Bread', 'Eggs', 'Butter', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Trash Cans', 'Bread', 'Syrup'}, {'Diapers', 'Milk', 'Water', 'Bread', 'Butter', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Diapers', 'Milk', 'Cola', 'Bread', 'Butter', 'Eggs', 'Trash Cans', 'Syrup'}, {'Water', 'Bread', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Milk', 'Bread', 'Butter', 'Eggs', 'Trash Cans', 'Tea'}, {'Eggs', 'Diapers'}, {'Milk', 'Bread', 'Eggs', 'Trash Cans', 'Tea'}, {'Pancake Mix', 'Milk', 'Cola', 'Diapers', 'Water', 'Trash Cans'}, {'Pancake Mix', 'Milk', 'Water', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Cola', 'Water', 'Bread', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Bread', 'Eggs', 'Butter', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Milk', 'Cola', 'Water', 'Bread', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Diapers', 'Cola', 'Milk', 'Bread', 'Butter', 'Trash Cans'}, {'Pancake Mix', 'Milk', 'Cola', 'Bread', 'Butter', 'Tea', 'Syrup'}, {'Pancake Mix', 'Milk', 'Cola', 'Diapers', 'Water', 'Bread', 'Trash Cans', 'Tea', 'Syrup'}, {'Butter', 'Tea', 'Trash Cans'}, {'Pancake Mix', 'Diapers', 'Milk', 'Water', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Milk', 'Bread', 'Eggs', 'Butter', 'Tea', 'Syrup'}, {'Diapers', 'Cola', 'Bread', 'Butter', 'Trash Cans', 'Tea'}, {'Pancake Mix', 'Diapers', 'Milk', 'Eggs', 'Butter'}, {'Cola', 'Tea', 'Syrup'}, {'Diapers', 'Cola', 'Water', 'Bread', 'Butter'}, {'Pancake Mix', 'Milk', 'Water', 'Bread', 'Eggs', 'Butter', 'Trash Cans', 'Syrup'}, {'Diapers', 'Milk', 'Cola', 'Water', 'Bread', 'Butter', 'Trash Cans'}, {'Pancake Mix', 'Diapers', 'Cola', 'Water', 'Bread', 'Eggs', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Water', 'Bread', 'Eggs', 'Tea'}, {'Butter', 'Trash Cans', 'Syrup'}, {'Milk', 'Bread'}, {'Pancake Mix', 'Diapers', 'Cola', 'Eggs', 'Butter', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Water', 'Bread', 'Butter', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Trash Cans', 'Diapers', 'Cola'}, {'Pancake Mix', 'Cola', 'Bread', 'Eggs', 'Butter', 'Trash Cans', 'Syrup'}, {'Milk', 'Water', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Milk', 'Diapers', 'Cola', 'Water', 'Trash Cans', 'Tea', 'Syrup'}, {'Trash Cans', 'Cola', 'Butter'}, {'Eggs', 'Cola', 'Tea'}, {'Diapers', 'Water', 'Bread', 'Eggs', 'Tea'}, {'Pancake Mix', 'Diapers', 'Cola', 'Water', 'Bread', 'Butter', 'Eggs', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Milk', 'Cola', 'Bread', 'Butter', 'Eggs'}, {'Milk', 'Cola', 'Diapers', 'Bread', 'Butter', 'Eggs', 'Trash Cans', 'Syrup'}, {'Milk', 'Water', 'Syrup'}, {'Diapers', 'Cola', 'Milk', 'Water', 'Bread', 'Eggs', 'Butter', 'Trash Cans', 'Tea'}, {'Pancake Mix', 'Diapers', 'Cola', 'Water', 'Eggs', 'Butter', 'Trash Cans', 'Syrup'}, {'Diapers', 'Milk', 'Cola'}, {'Eggs', 'Butter', 'Trash Cans', 'Syrup'}, {'Milk', 'Cola', 'Bread', 'Butter', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Butter', 'Cola', 'Bread'}, {'Pancake Mix', 'Milk', 'Water', 'Eggs', 'Butter', 'Trash Cans'}, {'Trash Cans', 'Water'}, {'Diapers', 'Water', 'Eggs', 'Butter', 'Trash Cans'}, {'Eggs', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Milk', 'Butter', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Bread', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Trash Cans', 'Diapers', 'Tea'}, {'Pancake Mix', 'Diapers', 'Milk', 'Eggs', 'Butter', 'Syrup'}, {'Pancake Mix', 'Butter', 'Diapers', 'Water'}, {'Milk', 'Diapers', 'Water', 'Bread', 'Eggs', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Eggs', 'Syrup'}, {'Diapers', 'Cola', 'Water', 'Trash Cans', 'Tea'}, {'Milk', 'Water', 'Bread', 'Eggs', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Bread', 'Eggs', 'Trash Cans', 'Tea'}, {'Pancake Mix', 'Butter', 'Tea', 'Syrup'}, {'Pancake Mix', 'Milk', 'Cola', 'Butter', 'Trash Cans'}, {'Pancake Mix', 'Milk', 'Cola', 'Bread', 'Butter', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Water', 'Tea', 'Syrup'}, {'Butter', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Milk', 'Cola', 'Water', 'Trash Cans', 'Tea', 'Syrup'}, {'Tea', 'Syrup'}, {'Pancake Mix', 'Milk', 'Cola', 'Bread', 'Butter', 'Trash Cans', 'Tea'}, {'Milk', 'Cola', 'Diapers', 'Water', 'Bread', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Butter', 'Water', 'Cola', 'Syrup'}, {'Milk', 'Diapers', 'Cola', 'Water', 'Bread', 'Butter', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Tea'}, {'Pancake Mix', 'Milk', 'Cola', 'Water', 'Bread', 'Butter', 'Eggs', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Water', 'Eggs', 'Butter', 'Trash Cans', 'Tea'}, {'Pancake Mix', 'Diapers', 'Cola', 'Water', 'Bread', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Eggs', 'Butter', 'Diapers', 'Tea'}, {'Pancake Mix', 'Tea'}, {'Milk', 'Cola', 'Diapers', 'Water', 'Bread', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Eggs', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Cola', 'Bread', 'Eggs', 'Butter', 'Tea', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Water', 'Tea', 'Syrup'}, {'Pancake Mix', 'Trash Cans', 'Cola', 'Butter'}, {'Diapers', 'Tea'}, {'Pancake Mix', 'Milk', 'Diapers', 'Water', 'Butter', 'Eggs', 'Trash Cans', 'Tea'}, {'Butter', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Water', 'Bread', 'Butter', 'Syrup'}, {'Pancake Mix', 'Water', 'Eggs', 'Trash Cans', 'Tea'}, {'Milk', 'Water', 'Eggs', 'Trash Cans', 'Tea'}, {'Water', 'Bread'}, {'Pancake Mix', 'Diapers', 'Cola', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Bread'}, {'Pancake Mix', 'Diapers', 'Cola', 'Milk', 'Bread', 'Butter', 'Eggs', 'Tea', 'Syrup'}, {'Diapers', 'Cola', 'Milk', 'Water', 'Eggs', 'Butter', 'Trash Cans', 'Syrup'}, {'Milk', 'Water', 'Butter', 'Trash Cans', 'Tea'}, {'Pancake Mix', 'Diapers', 'Cola', 'Water', 'Butter', 'Tea'}, {'Diapers', 'Milk', 'Bread'}, {'Pancake Mix', 'Cola', 'Water', 'Butter', 'Syrup'}, {'Cola', 'Diapers', 'Tea', 'Syrup'}, {'Milk', 'Tea'}, {'Diapers', 'Milk'}, {'Diapers', 'Cola', 'Bread'}, {'Pancake Mix', 'Diapers', 'Cola', 'Bread', 'Eggs', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Diapers', 'Milk', 'Cola', 'Water', 'Butter', 'Tea', 'Syrup'}, {'Water', 'Bread'}, {'Diapers', 'Cola', 'Bread', 'Eggs', 'Trash Cans'}, {'Pancake Mix', 'Diapers', 'Milk', 'Water', 'Bread', 'Syrup'}, {'Eggs', 'Cola', 'Butter'}, {'Diapers', 'Milk', 'Cola', 'Water', 'Eggs', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Bread', 'Tea', 'Syrup'}, {'Diapers', 'Bread', 'Eggs', 'Butter', 'Trash Cans'}, {'Pancake Mix', 'Milk', 'Diapers', 'Bread', 'Butter', 'Trash Cans'}, {'Butter', 'Milk', 'Cola'}, {'Pancake Mix', 'Diapers', 'Cola', 'Bread', 'Butter', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Cola', 'Milk', 'Bread', 'Butter', 'Eggs', 'Syrup'}, {'Eggs', 'Syrup', 'Bread'}, {'Pancake Mix', 'Diapers', 'Milk', 'Cola', 'Bread', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Diapers', 'Bread', 'Butter', 'Trash Cans'}, {'Pancake Mix', 'Milk', 'Cola', 'Diapers', 'Eggs', 'Butter', 'Tea', 'Syrup'}, {'Pancake Mix', 'Milk', 'Cola', 'Water', 'Bread', 'Butter', 'Eggs', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Milk', 'Water', 'Bread', 'Eggs', 'Trash Cans'}, {'Diapers', 'Cola', 'Milk', 'Eggs', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Pancake Mix', 'Milk', 'Bread', 'Eggs', 'Butter', 'Tea'}, {'Pancake Mix', 'Diapers', 'Bread', 'Butter', 'Trash Cans'}, {'Pancake Mix', 'Diapers', 'Cola', 'Milk', 'Bread', 'Butter', 'Syrup'}, {'Pancake Mix', 'Milk', 'Water', 'Bread', 'Eggs', 'Tea', 'Syrup'}, {'Diapers', 'Water', 'Bread'}, {'Butter', 'Syrup', 'Bread'}, {'Eggs', 'Butter', 'Syrup'}, {'Eggs', 'Diapers', 'Cola'}, {'Water', 'Cola'}, {'Pancake Mix', 'Trash Cans', 'Bread', 'Syrup'}, {'Eggs', 'Cola', 'Syrup', 'Bread'}, {'Pancake Mix', 'Milk', 'Butter', 'Trash Cans', 'Syrup'}, {'Pancake Mix', 'Butter', 'Cola', 'Bread'}, {'Pancake Mix', 'Milk', 'Cola', 'Diapers', 'Eggs', 'Butter', 'Trash Cans', 'Tea', 'Syrup'}, {'Milk', 'Cola', 'Eggs', 'Butter', 'Syrup'}, {'Eggs', 'Trash Cans'}, {'Pancake Mix', 'Trash Cans'}, {'Pancake Mix', 'Diapers', 'Cola', 'Milk', 'Eggs', 'Trash Cans', 'Tea', 'Syrup'}, {'Eggs', 'Diapers', 'Cola'}
        ]

print(len(transactions))
# Support function
def support(itemset, transactions):
    count = sum(1 for transaction in transactions if set(itemset).issubset(set(transaction)))
    return count / len(transactions)

# Confidence function
def confidence(lhs, rhs, transactions):
    lhs_support = support(lhs, transactions)
    if lhs_support == 0:
        return 0
    return support(set(lhs).union(set(rhs)), transactions) / lhs_support

# Function to generate candidate itemsets of a given size
def generate_candidates(itemset_list, size):
    return list(combinations(itemset_list, size))

# Apriori algorithm
def apriori(transactions, min_support, min_confidence):
    itemset = set(item for transaction in transactions for item in transaction)
    
    freq_itemsets = []
    itemsets_size = 1
    current_itemsets = generate_candidates(itemset, itemsets_size)
    
    while current_itemsets:
        valid_itemsets = [itemset for itemset in current_itemsets if support(itemset, transactions) >= min_support]
        freq_itemsets.extend(valid_itemsets)
        itemsets_size += 1
        current_itemsets = generate_candidates(itemset, itemsets_size)
    
    freq_itemsets = [itemset for itemset in freq_itemsets if len(itemset) > 1]

    rules = []
    for itemset in freq_itemsets:
        for i in range(1, len(itemset)):
            for lhs in combinations(itemset, i):
                rhs = set(itemset) - set(lhs)
                conf = confidence(lhs, rhs, transactions)
                if conf >= min_confidence:
                    rules.append((lhs, rhs, conf))
    
    return freq_itemsets, rules

# Number of transactions to modify (Delta)
def compute_delta(lhs, rhs, transactions, MCT):
    support_lhs_rhs = support(set(lhs).union(set(rhs)), transactions)
    support_rhs = support(rhs, transactions)
    # Ensure delta is large enough to reduce the support of sensitive rules
    delta = max(1, (support_lhs_rhs - (MCT * support_rhs)) * len(transactions))
    return delta

def item_frequencies(transactions):
    freq = {}
    for transaction in transactions:
        for item in transaction:
            if item in freq:
                freq[item] += 1
            else:
                freq[item] = 1
    return freq

# Function to calculate the accuracy based on the formula
def calculate_accuracy(original_db, modified_db):
    original_freq = item_frequencies(original_db)
    modified_freq = item_frequencies(modified_db)
    
    sum_original_freq = sum(original_freq.values())
    sum_modified_freq = sum(modified_freq.values())
    
    accuracy = 1 - ((sum_original_freq - sum_modified_freq) / sum_original_freq)
    
    return accuracy * 100  # Convert to percentage


# Fitness function
def fitness(selected_transactions, DB_prime, non_sensitive_rules, sensitive_rule, delta):
    modified_db = [t for t in DB_prime if t not in selected_transactions]
    
    lhs, rhs, _ = sensitive_rule
    sensitive_support = support(set(lhs).union(set(rhs)), modified_db)

    # Penalize heavily if sensitive support is above delta
    if sensitive_support >= delta:
        return float('-inf')
    
    valid_rules = 0
    for lhs, rhs, _ in non_sensitive_rules:
        if support(set(lhs).union(set(rhs)), modified_db) > 0:
            valid_rules += 1

    return valid_rules


# Genetic algorithm implementation
def genetic_algorithm(DB_prime, candidate_transactions, delta, non_sensitive_rules, sensitive_rule, pop_size=20, generations=100, mutation_rate=0.05):
    population = [random.sample(candidate_transactions, min(len(candidate_transactions), int(delta))) for _ in range(pop_size)]
    
    for generation in range(generations):
        population_fitness = [(individual, fitness(individual, DB_prime, non_sensitive_rules, sensitive_rule, delta)) for individual in population]
        population_fitness = [ind_fit for ind_fit in population_fitness if ind_fit[1] != float('-inf')]

        if not population_fitness:
            break

        population_fitness.sort(key=lambda x: x[1], reverse=True)

        parents = [individual for individual, _ in population_fitness[:pop_size // 2]]
        offspring = []
        for i in range(0, len(parents), 2):
            if i + 1 < len(parents):
                child1, child2 = crossover(parents[i], parents[i+1])
                offspring.extend([child1, child2])

        offspring = [mutate(child, candidate_transactions, mutation_rate) for child in offspring]
        population = parents + offspring

    best_individual = max(population, key=lambda individual: fitness(individual, DB_prime, non_sensitive_rules, sensitive_rule, delta))
    return best_individual

# Crossover function
def crossover(parent1, parent2):
    crossover_point = len(parent1) // 2
    child1 = parent1[:crossover_point] + parent2[crossover_point:]
    child2 = parent2[:crossover_point] + parent1[crossover_point:]
    return child1, child2

# Mutation function
def mutate(individual, candidate_transactions, mutation_rate):
    if random.random() < mutation_rate:
        index = random.randint(0, len(individual) - 1)
        new_transaction = random.choice(candidate_transactions)
        individual[index] = new_transaction
    return individual

# Function to calculate the utility based on the number of association rules
def calculate_utility(original_rules, modified_rules):
    ar_original = len(original_rules)
    ar_modified = len(modified_rules)
    
    utility = 1 - ((ar_original - ar_modified) / ar_original)
    
    return utility * 100  # Convert to percentage

# Main algorithm
def modify_database(transactions, rules, num_sensitive_rules=2, MST=0.4, MCT=0.6):
    DB_prime = transactions.copy()

    sorted_rules = sorted(rules, key=lambda r: r[2], reverse=True)
    sensitive_rules = sorted_rules[:num_sensitive_rules]

    SARcounter = 0
    non_sensitive_rules = [rule for rule in rules if rule not in sensitive_rules]

    start_time = time.time()
    while SARcounter < len(sensitive_rules):
        SARcounter += 1
        SAR = sensitive_rules[SARcounter - 1]

        rhs_items = list(SAR[1])
        item_supports = {item: support({item}, DB_prime) for item in rhs_items}
        victim_item = min(item_supports, key=item_supports.get)

        delta = compute_delta(SAR[0], SAR[1], DB_prime, MCT)
        candidate_transactions = [t for t in DB_prime if victim_item in t]

        selected_transactions = genetic_algorithm(DB_prime, candidate_transactions, delta, non_sensitive_rules, SAR)
        DB_prime = [t for t in DB_prime if t not in selected_transactions]

    num_lost_rules = len([rule for rule in non_sensitive_rules if support(set(rule[0]).union(set(rule[1])), DB_prime) == 0])

    end_time = time.time()
    cpu_time = end_time - start_time
    accuracy = calculate_accuracy(transactions, DB_prime)

# Re-run Apriori to get the new rules from the modified database (DB_prime)
    _, modified_rules = apriori(DB_prime, min_support, min_confidence)

    utility = calculate_utility(rules, modified_rules)  # Use the new utility calculation

    lost_rules_ratio = num_lost_rules / len(non_sensitive_rules)
    confidence_threshold_percent = MCT * 100

    # Calculate Hiding Failure with abs()
    SAR_in_D_prime = [rule for rule in sensitive_rules if support(set(rule[0]).union(set(rule[1])), DB_prime) > 0]
    hiding_failure = abs(len(SAR_in_D_prime)) / abs(len(sensitive_rules)) if sensitive_rules else 0 
    
    # Append results to the Excel file
    output_to_excel(cpu_time, accuracy, utility, lost_rules_ratio, confidence_threshold_percent, num_sensitive_rules, hiding_failure)

    return DB_prime

# Function to output results to Excel without deleting existing content
def output_to_excel(cpu_time, accuracy, utility, lost_rules_ratio, confidence_threshold_percent, num_sensitive_rules, hiding_failure):
    file_path = r'C:\Users\coola\OneDrive\Desktop\results.xlsx'  # Replace with your actual file path

    # Load the existing workbook and select the active worksheet
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Check if the sheet is empty; if so, add headers
    if sheet.max_row == 1:
        headers = ["CPU TIME (SEC)", "Accuracy %", "Utility%", "Lost Rules Ratio", "Confidence Threshold", "Number of Sensitive Rules", "Hiding Failure"]
        sheet.append(headers)

    # Append the results, including Hiding Failure
    row = [cpu_time, accuracy, utility, lost_rules_ratio, confidence_threshold_percent, num_sensitive_rules, hiding_failure]
    sheet.append(row)

    # Save the updated workbook
    workbook.save(file_path)

# Run the Apriori algorithm
min_support = 0.2
min_confidence = 0.6
freq_itemsets, rules = apriori(transactions, min_support, min_confidence)

# Modify the database
modified_db = modify_database(transactions, rules)
print(len(modified_db))

# Assuming modified_db is your modified list of transactions
modified_db = transactions  # Replace this with your actual modified_db after modification

# Specify the path using one of the options above
output_file_path = r"C:\Users\coola\OneDrive\Desktop\modified_transactions.txt"

# Write the modified_db transactions to the file
with open(output_file_path, 'w') as file:
    for transaction in modified_db:
        file.write(', '.join(transaction) + '\n')

print(f"Modified transactions written to {output_file_path}")


